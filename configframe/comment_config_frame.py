import os

from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import *
import openpyxl
from PyQt5.QtCore import Qt, QEvent
from config.root_directory import ROOT_DIR


class CommentWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # 设置窗口标题
        self.setWindowTitle('评论配置')
        self.width = 1002
        self.height = 642
        self.resize(self.width, self.height)

        self.sheet_names = ["娱乐", "带货", "情感", "其他"]

        # 创建按钮
        self.buttons = {}
        self.btn_1 = self.createButton(self.sheet_names[0],self.button1)
        self.btn_2 = self.createButton(self.sheet_names[1],self.button2)
        self.btn_3 = self.createButton(self.sheet_names[2],self.button3)
        self.btn_4 = self.createButton(self.sheet_names[3],self.button4)

        # 默认点击的按钮
        self.current_button = self.btn_1
        self.current_button.setStyleSheet('border: none; background-color: #E1E1E1; color: #C3473A;')

        # 选项卡页面
        self.tab1 = self.ui(self.sheet_names[0])
        self.tab2 = self.ui(self.sheet_names[1])
        self.tab3 = self.ui(self.sheet_names[2])
        self.tab4 = self.ui(self.sheet_names[3])

        self.initUI()

    def createButton(self, text, button_event):
        """
        创建按钮 （包括样式）
        :param text:
        :param button_event:
        :return:
        """
        button = QPushButton(text, self)
        font = QFont()
        font.setPointSize(14)
        button.setFont(font)
        # 设置按钮高度
        button.setFixedHeight(30)
        # 去边框 加背景
        button.setStyleSheet('border: none; background-color: #EDEDED; color: #2E2E2E;')
        # 禁用按钮默认设置
        # button.setAutoDefault(False)
        # 绑定按钮事件
        button.clicked.connect(button_event)
        # 绑定监听事件
        button.installEventFilter(self)
        self.buttons[text] = button  # 将按钮存储到字典中
        return button

    def eventFilter(self, obj, event):
        """
        监听鼠标事件（悬浮、离开、点击）
        :param obj:
        :param event:
        :return:
        """
        if obj in self.buttons.values():
            if event.type() == QEvent.Enter:
                if obj != self.current_button:
                    obj.setStyleSheet('border: none; background-color: #E6E6E6; color: #2E2E2E;')
            elif event.type() == QEvent.Leave:
                if obj != self.current_button:
                    obj.setStyleSheet('border: none; background-color: #EDEDED; color: #2E2E2E;')
            elif event.type() == QEvent.MouseButtonPress:
                self.current_button = obj
                obj.setStyleSheet('border: none; background-color: #E1E1E1; color: #C3473A;')
                for i in self.buttons.values():
                    if i != obj:
                        i.setStyleSheet('border: none; background-color: #EDEDED; color: #2E2E2E;')

        return super().eventFilter(obj, event)

    def initUI(self):
        # 创建一个垂直布局管理器（left_layout），用于放置侧边栏按钮和保存按钮
        left_layout = QVBoxLayout()
        # 向侧边栏添加一个小部件（按钮1）
        left_layout.addWidget(self.btn_1)
        left_layout.addWidget(self.btn_2)
        left_layout.addWidget(self.btn_3)
        left_layout.addWidget(self.btn_4)
        # 添加一个弹性空间以使按钮在布局中垂直居中
        left_layout.addStretch(5)
        # 设置按钮之间的垂直间距
        left_layout.setSpacing(0)

        # 创建容器
        left_widget = QWidget()
        # 设置容器内部由（left_layout）布局
        left_widget.setLayout(left_layout)
        # 设置容器的上、下、左、右边距都为零，以确保按钮紧贴布局的边缘
        left_layout.setContentsMargins(0, 0, 0, 0)

        # 创建一个可以具有多个选项卡的小部件
        self.right_widget = QTabWidget()
        # 设置选项卡的对象名称
        self.right_widget.tabBar().setObjectName("mainTab")
        # 向选项卡部件添加四个选项卡
        self.right_widget.addTab(self.tab1, '')
        self.right_widget.addTab(self.tab2, '')
        self.right_widget.addTab(self.tab3, '')
        self.right_widget.addTab(self.tab4, '')
        # 切换选项卡页（设置默认为第一个选项卡）
        self.right_widget.setCurrentIndex(0)
        # 自定义选项卡样式（此处样式为：隐藏选项栏、去边框）
        self.right_widget.setStyleSheet('''
            QTabBar::tab{width: 0; height: 0; margin: 0; padding: 0; border: none;}
            QTabWidget::pane { border: none; }
        ''')

        # 创建一个水平布局管理器（main_layout）用于在水平方向上从左到右排列小部件
        main_layout = QHBoxLayout()
        # 将 左侧 按钮容器小部件添加到布局中
        main_layout.addWidget(left_widget)
        # 将 右侧 选项卡小部件添加到布局中
        main_layout.addWidget(self.right_widget)
        # 设置左侧部件在布局中的伸缩性，索引为0，占比为50
        main_layout.setStretch(0, 50)
        # 设置右侧部件在布局中的伸缩性，索引为1，占比为200
        main_layout.setStretch(1, 200)
        # 设置内部的小部件与布局边缘的距离（左上右下）
        main_layout.setContentsMargins(0,0,0,0)

        # 创建一个容器（用于容纳整体布局）
        main_widget = QWidget()
        # 设置容器内部由（main_layout）布局
        main_widget.setLayout(main_layout)

        # 将该容器设置为主窗口的中央部件
        self.setCentralWidget(main_widget)

    # 按钮点击事件
    def button1(self):
        self.right_widget.setCurrentIndex(0)

    def button2(self):
        self.right_widget.setCurrentIndex(1)

    def button3(self):
        self.right_widget.setCurrentIndex(2)

    def button4(self):
        self.right_widget.setCurrentIndex(3)

    # 选项卡内容
    def ui(self,sheet_name):
        table_widget = self.create_table_widget(sheet_name)
        return table_widget

    def create_table_widget(self, sheet_name):
        """
        创建表格部件
        :param sheet_name:
        :return: 容器小部件
        """
        # 创建一个表格部件
        table_widget = QTableWidget()
        # 设置选中项的背景颜色为绿色
        table_widget.setStyleSheet("QTableWidget::item:selected { background-color: #418F1F; }")
        # 加载数据到表格
        self.load_sheet(sheet_name, table_widget)

        # # 创建一个垂直布局管理器（用于控制小部件的排列和布局）
        # layout = QVBoxLayout()
        # # 将表格部件放入垂直布局管理器
        # layout.addWidget(table_widget)
        # # 创建一个容器
        # container_widget = QWidget()
        # # 容器内部由（layout）布局
        # container_widget.setLayout(layout)
        # # # 设置容器的上、下、左、右边距
        # # layout.setContentsMargins(0, 0, 0, 0)
        # # 设置容器小部件的高度
        # container_widget.setFixedHeight(450)

        # 当表格中的项目（单元格）发生更改时，会触发 itemChanged 信号，并执行自动保存
        table_widget.itemChanged.connect(self.auto_save_data)
        # 为表格部件绑定右键点击事件
        table_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        table_widget.customContextMenuRequested.connect(self.showContextMenu)

        return table_widget

    def add_tail_blank_row(self,table_widget,sheet):
        """
        添加尾部空白行
        :param table_widget: 要填充数据的Qt表格部件
        :param sheet: 要加载的Excel工作表
        :return: 最大行、最大列
        """
        # 获取工作表的总行数
        max_rows = sheet.max_row
        max_column = sheet.max_column

        # 检查最后一行的单元格是否有一个不为空
        last_row_not_empty = any(sheet.cell(row=max_rows, column=col).value for col in
                                 range(1, max_column + 1) if sheet.cell(row=max_rows, column=col).value)
        # 如果最后一行的单元格有一个不为空
        if last_row_not_empty:
            # 添加一行空白单元格
            table_widget.insertRow(max_rows - 1)
            max_rows += 1
        return max_rows, max_column

    def showContextMenu(self, pos):
        """
        显示上下文菜单
        :param table_widget: 当前的Qt表格部件
        :param pos: 鼠标点击位置
        """
        # 获取当前选中的选项卡对应的小部件
        table_widget = self.right_widget.currentWidget()
        if isinstance(table_widget,QTableWidget):
            # 获取选中的单元格
            selected_items = table_widget.selectedItems()

            # 如果没有选中的单元格，则不显示右键菜单
            if not selected_items:
                return

        # 创建右键上下文菜单
        contextMenu = QMenu(self)
        deleteAction = QAction('删除', self)
        deleteAction.triggered.connect(self.deleteSelectedCells)
        contextMenu.addAction(deleteAction)

        # 显示右键上下文菜单在鼠标点击位置
        contextMenu.exec_(table_widget.mapToGlobal(pos))

    def deleteSelectedCells(self):
        # 获取当前选中的选项卡对应的小部件
        table_widget = self.right_widget.currentWidget()
        if isinstance(table_widget, QTableWidget):
            # 删除选中的单元格
            for item in table_widget.selectedItems():
                table_widget.setItem(item.row(), item.column(), QTableWidgetItem(''))

    def keyPressEvent(self, event):
        """
        重写 keyPressEvent 方法来捕获和处理键盘事件
        :param event:
        :return:
        """
        # 捕获键盘的 del 按钮
        if event.key() == Qt.Key_Backspace:
            # 当按下 Delete 键时触发删除操作
            self.deleteSelectedCells()
        else:
            super().keyPressEvent(event)

    def load_sheet(self, sheet_name, table_widget):
        """
        从Excel加载数据填充表格
        :param sheet_name: 要加载的Excel工作表的名称
        :param table_widget: 要填充数据的Qt表格部件
        """
        excel_file = os.path.join(ROOT_DIR, 'config', 'comment_data.xlsx')
        # 加载Excel数据
        wb = openpyxl.load_workbook(excel_file)
        # 选择指定的工作表
        sheet = wb[sheet_name]

        # 获取工作表的列名称作为表格的列标题
        column_names = [str(sheet.cell(row=1, column=col).value) for col in range(1, len(self.sheet_names)+1)]

        # 设置表格的列数和列标题
        table_widget.setColumnCount(len(column_names))
        table_widget.setHorizontalHeaderLabels(column_names)

        # 获取列标题对象
        header = table_widget.horizontalHeader()
        # 设置列标题对象的高度为40像素
        header.setStyleSheet("QHeaderView::section { height: 25px; }")
        # 设置自动拉伸表格（Interactive：允许用户手动调整列和宽；Stretch：列均匀拉伸；ResizeToContents：根据内容自动调整列宽度）
        header.setSectionResizeMode(QHeaderView.Stretch)
        # 创建一个字体对象
        font = QFont()
        font.setPointSize(14)  # 设置字体大小为12像素
        # 将字体应用于水平表头
        header.setFont(font)

        # 添加尾部空白行
        max_rows, max_column = self.add_tail_blank_row(table_widget,sheet)

        # 设置表格的行数（减去1是因为第一行是标题行）
        table_widget.setRowCount(max_rows - 1)
        # 遍历表中的数据
        for row in range(2, max_rows + 1):  # 从第二行开始，因为第一行是标题
            for col in range(1, max_column + 1):  # 遍历每一列
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is None:  # 将为None的单元格替换为空字符串
                    cell_value = ""
                # 创建一个单元格，并设置单元格数据
                item = QTableWidgetItem(str(cell_value))
                # 将上述单元格添加到表格部件指定的行和列
                table_widget.setItem(row - 2, col - 1, item)  # 减去2是因为表格行索引从0开始

    # 自动保存数据到Excel
    def auto_save_data(self):
        # 获取当前选项卡的索引
        current_tab_index = self.right_widget.currentIndex()
        sheet_name = self.sheet_names[current_tab_index]
        # 获取当前选中的选项卡对应的小部件
        table_widget = self.right_widget.currentWidget()

        if isinstance(table_widget,QTableWidget):
            excel_file = os.path.join(ROOT_DIR,"config","comment_data.xlsx")
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb[sheet_name]

            # 清空工作表内容
            sheet.delete_rows(2, sheet.max_row)  # 删除第二行及以后的内容，保留第一行表头

            # 将数据保存到 Excel 中
            for row in range(table_widget.rowCount()):
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    if item:
                        sheet.cell(row=row + 2, column=col + 1, value=item.text())

            # 添加尾部空白行
            self.add_tail_blank_row(table_widget,sheet)

            # 保存 Excel 文件
            wb.save(excel_file)
            QMessageBox.information(self, '提示', '数据已自动保存')
        else:
            QMessageBox.warning(self, '警告', '未找到数据表格')


if __name__ == '__main__':
    app = QApplication([])
    ex = CommentWindow()
    ex.show()
    app.exec_()
